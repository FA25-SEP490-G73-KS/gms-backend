#!/usr/bin/env python3
"""
Convert CSV file to Excel format.
"""

import csv
import sys

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")
    print("Will create a simple Excel-compatible file instead.")

def csv_to_excel(csv_file, excel_file):
    """Convert CSV to Excel."""
    
    if HAS_OPENPYXL:
        # Use openpyxl for proper Excel format
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Case Matrices"
        
        # Read CSV and write to Excel
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_file)
        print(f"Excel file created: {excel_file}")
    else:
        # Fallback: Create a tab-separated file that Excel can open
        with open(csv_file, 'r', encoding='utf-8') as f_in:
            with open(excel_file.replace('.xlsx', '.tsv'), 'w', encoding='utf-8') as f_out:
                reader = csv.reader(f_in)
                writer = csv.writer(f_out, delimiter='\t')
                for row in reader:
                    writer.writerow(row)
        print(f"Tab-separated file created: {excel_file.replace('.xlsx', '.tsv')}")
        print("   (You can open this file in Excel)")

if __name__ == "__main__":
    csv_file = "test_case_matrices_8_20.csv"
    excel_file = "test_case_matrices_8_20.xlsx"
    
    csv_to_excel(csv_file, excel_file)

